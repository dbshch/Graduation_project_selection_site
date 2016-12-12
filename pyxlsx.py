# coding=UTF-8
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import datetime


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border += top
    for cell in rows[-1]:
        cell.border += bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border += left
        r.border += right
        if fill:
            for c in row:
                c.fill = fill


def multiCell(ws, value, rng, b=False, color="000000", horizontal="center", vertical="center"):
    my_cell = ws[rng.split(":")[0]]
    my_cell.value = value
    font = Font(b=b, color=color)
    al = Alignment(horizontal=horizontal, vertical=vertical)
    style_range(ws, rng, font=font, alignment=al)


def export(data, output):
    wb = Workbook()
    ws = wb.active
    multiCell(ws, "Proj. No.", "A1:A2")
    multiCell(ws, "Title", "B1:B2")
    multiCell(ws, "Preference", "C1:E1")
    for i in range(3):
        cell = ws["%s2" % chr(ord('C') + i)]
        cell.value = i + 1
        cell.alignment = Alignment(horizontal="center", vertical="center")
    end_of_project_cell = 2
    end_of_group_cell = [2, 2, 2]
    for project in data:
        if project['groups'][0] or project['groups'][1] or project['groups'][2]:
            for i in range(3):
                pref = project['groups'][i]
                if pref:
                    for group in pref:
                        for student in group:
                            end_of_group_cell[i] += 1
                            ws['%s%d' % (chr(67 + i), end_of_group_cell[i])] = student
                        end_of_group_cell[i] += 1
            tmp = max(end_of_group_cell) + 2
            multiCell(ws, project['i'], "A%d:A%d" % (end_of_project_cell + 1, tmp))
            multiCell(ws, project['title'], "B%d:B%d" % (end_of_project_cell + 1, tmp))
            for i in range(3):
                end_of_group_cell[i] = tmp
            end_of_project_cell = tmp

    wb.save(output)


if __name__ == "__main__":
    data = [{'i': 1, 'title': 'Freesense-Backlight Panel Defect Recognition', 'groups': [[['吴承刚']], [[]], [[]]]}]
    data.append({'i': 2, 'title': 'Recognition', 'groups': [[['钱神']], [['负心汉', '陈亦轩']], [[]]]})
    data.append({'i': 3, 'title': 'abcd', 'groups': [[['钱神']], [['负心汉', '陈亦轩']], [['钱神'], ['负心汉', '陈亦轩']]]})
    export(data, 'test.xlsx')
